import openpyxl

# Crear un libro de trabajo y una hoja
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Datos"

# Escribir datos en celdas
ws['A1'] = "Nombre"
ws['B1'] = "Edad"
ws['A2'] = "Juan"
ws['B2'] = 30
ws['A3'] = "Ana"
ws['B3'] = 25

# Guardar el archivo
wb.save("datos.xlsx")

# Leer datos de un archivo Excel
wb = openpyxl.load_workbook("datos.xlsx")
ws = wb.active
for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=2):
    for cell in row:
        print(cell.value)